import numpy as np
import openpyxl
import pandas as pd
from sklearn.metrics import mean_squared_error, mean_absolute_error


def get_original_data():
    """
    获取原始数据
    """
    # 全部结果
    path_origin = r'D:\Sustech\course\CS490_Graduation_Thesis\work\data-preprocessed\output.xlsx'
    origin = {}
    dx = pd.read_excel(path_origin, sheet_name='output')
    row = dx[dx.iloc[:, 0].str.contains('', na=False)]
    for i in range(len(row)):
        row_data = row.iloc[i]
        student_id = row_data[1]
        origin[student_id] = {
            '最终打分': row_data['最终得分'],
        }
    return origin


def get_model_results(path, folder, model):
    """
    获取模型预测结果
    """
    # read excel
    result = {}
    dx = pd.read_excel(rf"{path}\{model}\{folder}.xlsx", sheet_name=f'{model}')
    for i in range(len(dx)):
        row_data = dx.iloc[i]
        student_id = row_data[0]
        result[student_id] = {
            '最终打分': row_data['打分'],
        }
    # print(model_results)
    return result


def calculate_metrics(origin_data, result_data):
    """
    生成评估参数
    """
    # 计算MSE和MAE
    mse = mean_squared_error(origin_data, result_data)
    mae = mean_absolute_error(origin_data, result_data)

    # 计算Pearson相关系数（增加标准差校验）
    with np.errstate(divide='ignore', invalid='ignore'):
        cov_matrix = np.cov(origin_data, result_data)
        if cov_matrix[0, 0] == 0 or cov_matrix[1, 1] == 0:
            pcc = 0.0  # 处理零方差情况
        else:
            pcc = cov_matrix[0, 1] / np.sqrt(cov_matrix[0, 0] * cov_matrix[1, 1])

    return {
        'MSE': mse,
        'MAE': mae,
        'PCC': pcc if not np.isnan(pcc) else 0.0  # 二次容错
    }


def get_metrics(origin, result):
    """
    根据教师评分原始结果，以及模型打分结果，得到评估参数
    """
    # transform format to get metrics
    scores = {
        "学号": [],
        "最终打分": [[], []],
    }
    for key in result.keys():
        scores["学号"].append(key)
        for i in range(1, len(scores)):
            key_name = list(scores.keys())[i]
            scores[key_name][0].append(origin[key][key_name])
            scores[key_name][1].append(result[key][key_name])
    # calculate metrics
    metrics_results = {}
    for key in scores.keys():
        if key == '学号':
            continue
        metrics_result = calculate_metrics(scores[key][0], scores[key][1])
        for metric in metrics_result.keys():
            if metric not in metrics_results.keys():
                metrics_results[metric] = {}
            metrics_results[metric][key] = metrics_result[metric]
    # print(f"compare result:\n {compare_result}\n")
    return metrics_results


def get_average(data):
    """
    生成由均值
    """
    average_results = {}
    for SID in data.keys():
        score = data[SID]
        for key in score.keys():
            if average_results.get(key) is None:
                average_results[key] = []
            average_results[key].append(score[key])
    for key in average_results.keys():
        average_results[key] = np.mean(average_results[key])
    return average_results


def form_xlsx_metrics(all_results, input_path, folder_name):
    """
    生成由分数、系数组成的xlsx文件
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    for key in all_results.keys():
        ws = default_sheet if key == list(all_results.keys())[0] else wb.create_sheet()
        ws.title = "MSE(均方误差)" if key == 'MSE' else "MAE(平均绝对误差)" if key == 'MAE' else "PCC(皮尔逊相关系数)" if key == 'PCC' else "unknown"
        data = all_results[key]
        ws.append(["模型"] + list(data[list(data.keys())[0]].keys()))
        for model_name in data.keys():
            ws.append([model_name] + [data[model_name][score] for score in data[model_name].keys()])
    saving_path = rf"{input_path}\metrics_{folder_name}.xlsx"
    wb.save(saving_path)
    print(f"successfully save the {saving_path}")


def form_xlsx_average(all_results, input_path, folder_name):
    """
    生成由分数、系数组成的xlsx文件
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{folder_name}_平均值"
    ws.append(["模型"] + list(all_results[list(all_results.keys())[0]].keys()))
    for key in all_results.keys():
        ws.append([key] + [all_results[key][score] for score in all_results[key].keys()])
    saving_path = rf"{input_path}\average_{folder_name}.xlsx"
    wb.save(saving_path)
    print(f"successfully save the {saving_path}")


if __name__ == "__main__":
    path = rf"D:\Sustech\course\CS490_Graduation_Thesis\work\results\aliyunbailian"
    models = [
        # 通义 - 文本生成
        "通义千问-Plus",
        "通义千问2.5-14B-1M",
        "通义千问-Turbo",
        # DeepSeek - 文本生成
        "DeepSeek-V3",
        # DeepSeek - 推理模型
        "DeepSeek-R1",
    ]
    folder = "all_without_system"
    metrics = {}
    average = {}
    origin = get_original_data()
    average["教师评分结果"] = get_average(
        origin
        # {list(origin.keys())[i]: origin[list(origin.keys())[i]] for i in range(5)}
    )
    for model in models:
        print(f"start to calculate the result from {model}")
        # get data
        result = get_model_results(path, folder, model)
        average[model] = get_average(result)
        # get metrics
        metric_result = get_metrics(origin, result)
        for metric in metric_result.keys():
            if metric not in metrics.keys():
                metrics[metric] = {}
            metrics[metric][model] = metric_result[metric]
    form_xlsx_metrics(metrics, path, folder)
    form_xlsx_average(average, path, folder)
