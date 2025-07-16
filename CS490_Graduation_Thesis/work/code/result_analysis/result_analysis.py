import numpy as np
import openpyxl
import pandas as pd
from sklearn.metrics import mean_squared_error, mean_absolute_error


def get_original_data():
    """
    获取原始数据
    """
    # 全部结果
    path_origin = r'../../data-preprocessed/output.xlsx'
    origin = {}
    dx = pd.read_excel(path_origin, sheet_name='output')
    row = dx[dx.iloc[:, 0].str.contains('', na=False)]
    for i in range(len(row)):
        row_data = row.iloc[i]
        student_id = row_data[1]
        origin[student_id] = {
            '最终打分': row_data['最终得分'],
            '结构完整性得分': row_data['报告结构的完整性 20%'],
            '逻辑清晰度得分': row_data['报告逻辑的清晰度 20%'],
            '语言连贯性得分': row_data['报告语言的连贯性 20%'],
            '内容独特性和创新性得分': row_data['报告内容的独特性和创新性 20%'],
            '参考文献规范性得分': row_data['参考文献的规范性 10%'],
            '课程知识掌握度得分': row_data['课程知识的掌握程度 10%'],
        }
    return origin


def get_model_results(path, folder, model):
    """
    获取模型预测结果
    """
    # read excel
    result = {}
    dx = pd.read_excel(rf"{path}\{model}\{folder}.xlsx", sheet_name=f'{model}')
    for i in range(len(dx)):
        row_data = dx.iloc[i]
        student_id = row_data[0]
        result[student_id] = {
            '最终打分': row_data['最终打分'],
            '结构完整性得分': row_data['结构完整性得分'],
            '逻辑清晰度得分': row_data['逻辑清晰度得分'],
            '语言连贯性得分': row_data['语言连贯性得分'],
            '内容独特性和创新性得分': row_data['内容独特性和创新性得分'],
            '参考文献规范性得分': row_data['参考文献规范性得分'],
            '课程知识掌握度得分': row_data['课程知识掌握度得分'],
        }
    # print(model_results)
    return result


def calculate_metrics(origin_data, result_data):
    """
    生成评估参数
    """
    # 计算MSE和MAE
    mse = mean_squared_error(origin_data, result_data)
    mae = mean_absolute_error(origin_data, result_data)

    # 计算Pearson相关系数（增加标准差校验）
    with np.errstate(divide='ignore', invalid='ignore'):
        cov_matrix = np.cov(origin_data, result_data)
        if cov_matrix[0, 0] == 0 or cov_matrix[1, 1] == 0:
            pcc = 0.0  # 处理零方差情况
        else:
            pcc = cov_matrix[0, 1] / np.sqrt(cov_matrix[0, 0] * cov_matrix[1, 1])

    return {
        'MAE': round(mae, 2),
        'MSE': round(mse, 2),
        'PCC': round(pcc, 2) if not np.isnan(pcc) else 0.0  # 二次容错
    }


def get_metrics(origin, result):
    """
    根据教师评分原始结果，以及模型打分结果，得到评估参数
    """
    # transform format to get metrics
    scores = {
        "学号": [],
        "最终打分": [[], []],
        "结构完整性得分": [[], []],
        "逻辑清晰度得分": [[], []],
        "语言连贯性得分": [[], []],
        "内容独特性和创新性得分": [[], []],
        "参考文献规范性得分": [[], []],
        "课程知识掌握度得分": [[], []],
    }
    for key in result.keys():
        scores["学号"].append(key)
        for i in range(1, len(scores)):
            key_name = list(scores.keys())[i]
            scores[key_name][0].append(origin[key][key_name])
            scores[key_name][1].append(result[key][key_name])
    # calculate metrics
    metrics_results = {}
    for key in scores.keys():
        if key == '学号':
            continue
        metrics_result = calculate_metrics(scores[key][0], scores[key][1])
        for metric in metrics_result.keys():
            if metric not in metrics_results.keys():
                metrics_results[metric] = {}
            metrics_results[metric][key] = metrics_result[metric]
    # print(f"compare result:\n {compare_result}\n")
    return metrics_results


def get_average(data):
    """
    生成由均值
    """
    average_results = {}
    for SID in data.keys():
        score = data[SID]
        for key in score.keys():
            if average_results.get(key) is None:
                average_results[key] = []
            average_results[key].append(score[key])
    for key in average_results.keys():
        average_results[key] = round(np.mean(average_results[key]), 2)
    return average_results


def form_xlsx_metrics(all_results, input_path, folder_name):
    """
    生成由分数、系数组成的xlsx文件
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "最终打分"
    ws.append(["模型"] + list(all_results.keys()))
    for model_name in all_results[list(all_results.keys())[0]]:
        line = [model_name]
        for key in all_results.keys():
            line.append(all_results[key][model_name]["最终打分"])
        ws.append(line)
    for key in all_results.keys():
        ws = wb.create_sheet()
        ws.title = "MSE(均方误差)" if key == 'MSE' else "MAE(平均绝对误差)" if key == 'MAE' else "PCC(皮尔逊相关系数)" if key == 'PCC' else "unknown"
        data = all_results[key]
        ws.append(["模型"] + list(data[list(data.keys())[0]].keys()))
        for model_name in data.keys():
            ws.append([model_name] + [data[model_name][score] for score in data[model_name].keys()])
    saving_path = rf"{input_path}\metrics_{folder_name}.xlsx"
    wb.save(saving_path)
    print(f"successfully save the {saving_path}")


def form_xlsx_average(all_results, input_path, folder_name):
    """
    生成由分数、系数组成的xlsx文件
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "avg(平均值)"
    ws.append(["模型"] + list(all_results[list(all_results.keys())[0]].keys()))
    for key in all_results.keys():
        ws.append([key] + [all_results[key][score] for score in all_results[key].keys()])
    saving_path = rf"{input_path}\average_{folder_name}.xlsx"
    wb.save(saving_path)
    print(f"successfully save the {saving_path}")


if __name__ == "__main__":
    # 阿里云百炼 aliyunbailian
    aliyunbailian_path = rf"../../results/aliyunbailian"
    aliyunbailian_models = [
        # 通义 - 文本生成
        "通义千问-Plus",
        "通义千问2.5-14B-1M",
        "通义千问-Turbo",
        # DeepSeek - 文本生成
        "DeepSeek-V3",
        # DeepSeek - 推理模型
        "DeepSeek-R1",
        "通义千问-Turbo-Latest",
    ]
    # 网页端大语言模型 llmwebui
    llmwebui_path = rf"/results/LLM_web-ui"
    llmwebui_models = [
        "chatGPT",
        "deepseek",
        "tongyiqianwen",
        "wenxinyiyan",
    ]
    print("start to compare")
    folder_list = [
        'all_system-original_few-shot-0',
        'all_system-original_few-shot-1',
        'all_system-original_few-shot-2',
        'all_system-original_few-shot-3',
        'all_system-w-o-role_few-shot-2',
    ]
    path = aliyunbailian_path
    models = aliyunbailian_models
    metrics = {}
    average = {}
    for folder in folder_list:
        print(f"for {folder}")
        origin = get_original_data()
        average["教师评分结果"] = get_average(
            origin
            # {list(origin.keys())[i]: origin[list(origin.keys())[i]] for i in range(5)}
        )
        for model in models:
            print(f"start to calculate the result from {model}")
            # get data
            result = get_model_results(path, folder, model)
            average[model] = get_average(result)
            # get metrics
            metric_result = get_metrics(origin, result)
            for metric in metric_result.keys():
                if metric not in metrics.keys():
                    metrics[metric] = {}
                metrics[metric][model] = metric_result[metric]
        form_xlsx_metrics(metrics, path, folder)
        # form_xlsx_average(average, path, folder)
        print()
